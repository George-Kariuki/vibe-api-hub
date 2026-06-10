import csv
import io
from typing import Any

from fastapi import HTTPException

SUPPORTED_TYPES = {"csv", "xlsx"}


def detect_file_type(filename: str) -> str:
    """Return 'csv' or 'xlsx' based on the file extension, or raise 422."""
    if not filename or "." not in filename:
        raise HTTPException(
            status_code=422,
            detail="Could not determine file type. Upload a .csv or .xlsx file.",
        )
    ext = filename.rsplit(".", 1)[-1].lower()
    if ext not in SUPPORTED_TYPES:
        raise HTTPException(
            status_code=422,
            detail=f"Unsupported file type '.{ext}'. Only .csv and .xlsx are supported.",
        )
    return ext


def _build_keys(header_cells: list[Any]) -> list[str]:
    """Turn a header row into unique, non-empty string keys."""
    keys: list[str] = []
    seen: dict[str, int] = {}
    for i, cell in enumerate(header_cells):
        name = str(cell).strip() if cell is not None else ""
        if not name:
            name = f"column_{i + 1}"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
        keys.append(name)
    return keys


def _row_to_object(keys: list[str], values: list[Any]) -> dict[str, Any]:
    obj: dict[str, Any] = {}
    for i, key in enumerate(keys):
        obj[key] = values[i] if i < len(values) else None
    return obj


def _is_empty_row(values: list[Any]) -> bool:
    return all(v is None or (isinstance(v, str) and v.strip() == "") for v in values)


def _coerce_cell(value: Any) -> Any:
    """Coerce a cell value to a JSON-safe primitive."""
    if value is None:
        return None
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def parse_csv(
    content: bytes,
    has_header: bool,
    delimiter: str,
    skip_empty_rows: bool,
    max_rows: int,
) -> tuple[list[dict[str, Any]], list[str]]:
    if len(delimiter) != 1:
        raise HTTPException(
            status_code=422,
            detail="delimiter must be a single character, e.g. ',', ';', or '\\t'.",
        )

    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            text = content.decode("latin-1")
        except UnicodeDecodeError:
            raise HTTPException(
                status_code=422,
                detail="Could not decode the CSV file. Ensure it is UTF-8 or Latin-1 encoded.",
            )

    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    all_rows = [row for row in reader]

    if skip_empty_rows:
        all_rows = [r for r in all_rows if not _is_empty_row(r)]

    if not all_rows:
        raise HTTPException(status_code=422, detail="The file contains no parseable rows.")

    if has_header:
        keys = _build_keys(all_rows[0])
        data_rows = all_rows[1:]
    else:
        width = max((len(r) for r in all_rows), default=0)
        keys = [f"column_{i + 1}" for i in range(width)]
        data_rows = all_rows

    if len(data_rows) > max_rows:
        raise HTTPException(
            status_code=422,
            detail=f"File has {len(data_rows)} rows, which exceeds the max_rows limit of {max_rows}.",
        )

    rows = [_row_to_object(keys, [_coerce_cell(v) for v in r]) for r in data_rows]
    return rows, keys


def parse_xlsx(
    content: bytes,
    has_header: bool,
    sheet_index: int,
    skip_empty_rows: bool,
    max_rows: int,
) -> tuple[list[dict[str, Any]], list[str]]:
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(
            status_code=422,
            detail="Could not read the Excel file. Ensure it is a valid .xlsx file.",
        )

    try:
        sheet_names = workbook.sheetnames
        if sheet_index < 0 or sheet_index >= len(sheet_names):
            raise HTTPException(
                status_code=422,
                detail=(
                    f"sheet_index {sheet_index} is out of range. "
                    f"This file has {len(sheet_names)} sheet(s) (valid indices 0 to {len(sheet_names) - 1})."
                ),
            )
        worksheet = workbook[sheet_names[sheet_index]]

        all_rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
    finally:
        workbook.close()

    if skip_empty_rows:
        all_rows = [r for r in all_rows if not _is_empty_row(r)]

    if not all_rows:
        raise HTTPException(status_code=422, detail="The file contains no parseable rows.")

    if has_header:
        keys = _build_keys(all_rows[0])
        data_rows = all_rows[1:]
    else:
        width = max((len(r) for r in all_rows), default=0)
        keys = [f"column_{i + 1}" for i in range(width)]
        data_rows = all_rows

    if len(data_rows) > max_rows:
        raise HTTPException(
            status_code=422,
            detail=f"File has {len(data_rows)} rows, which exceeds the max_rows limit of {max_rows}.",
        )

    rows = [_row_to_object(keys, [_coerce_cell(v) for v in r]) for r in data_rows]
    return rows, keys
